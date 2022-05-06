import openpyxl
import getopt
import sys


def init(xlsx_path):
    workbook = openpyxl.load_workbook(xlsx_path)
    sheet = workbook['Sheet1']
    return sheet


# Function takes list of strings as input.
# Returns list of preprocessed strings (removes spaces, '(S)', and 'S') and a mapping to original string.
def process_strings(string_list):
    true_value_mapping = {}
    return_list = list()
    for i in range(len(string_list)):
        if string_list[i] == "":
            continue
        true_value = string_list[i]
        return_list.append(string_list[i].replace(" ", ""))
        if return_list[len(return_list) - 1][-3: len(return_list[len(return_list) - 1])] == "(S)":
            return_list[len(return_list) - 1] = return_list[len(return_list) - 1][0: len(return_list[len(return_list) -
                                                                                                     1]) - 3]
        elif return_list[len(return_list) - 1][-1] == 'S':
            return_list[len(return_list) - 1] = return_list[len(return_list) - 1][0: len(return_list[len(return_list) -
                                                                                                     1]) - 1]
        true_value_mapping.update({return_list[len(return_list) - 1]: true_value})
    return return_list, true_value_mapping


# Function takes cell contents of structured or unstructured annotation from query file as input.
# Returns list of preprocessed strings (removes spaces, '(S)', and 'S') and a mapping to original string.
def get_values(cell):
    if cell[0] == '(':
        cell = cell[1: len(cell)]
    if cell[-3:len(cell)] != "(S)" and cell[len(cell)-1] == ')':
        cell = cell[0: len(cell)-1]
    string_list = cell.split(',')
    string_list,  true_value_mapping = process_strings(string_list)
    return string_list, true_value_mapping


# Function returns the common strings (strict equality) from 2 string lists
def get_common_values(list1, list2):
    set1 = set(list1)
    set2 = set(list2)
    if set1 & set2:
        return list(set1 & set2)
    return list()


# Function takes a list of preprocessed strings and a mapping to its original string as inout.
# Returns string in a format suitable to be written to query file
def format_list(string_list, value_mapping):
    res = "("
    for i in string_list:
        res += value_mapping.get(i)
        res += ","
    res = res[0:len(res)-1]
    res += ")"
    return res


# Function creates a new annotated file for exact match strings
def annotate_exact_match(sheet, res_file, column_priority):
    results_workbook = openpyxl.Workbook()
    results_sheet = results_workbook.active
    i = 0
    for row in sheet.iter_rows(max_row=1):  # loop to create header row
        res_row = list()
        for cell in row:
            res_row.append(cell.value)
            i += 1
        res_row.append("Multimodal_Answer")
        results_sheet.append(res_row)
    count_match = 0  # count for rows that have matches between structured and unstructured
    count_populated_cells = 0  # count for rows with at least some value in structured and unstructured
    for row_no, row in enumerate(sheet.iter_rows(min_row=2), start=1):
        for j in range(len(row)):  # loop copies all columns from inout file to output file
            results_sheet.cell(row_no + 1, j + 1).value = row[j].value

        if row[i - 2].value is None:  # if structured cell is empty, print unstructured cell to result
            results_sheet.cell(row_no + 1, i + 1).value = row[i - 1].value
        elif row[i - 1].value is None:  # if unstructured cell is empty, print structured cell to result
            results_sheet.cell(row_no + 1, i + 1).value = row[i - 2].value
        else:
            count_populated_cells += 1
            if column_priority == 0:  # case for higher priority for structured data
                structured, structured_true_value_mapping = get_values(row[i - 2].value)  # get list of pre-processed prescriptions and a mapping to original string
                unstructured, _ = get_values(row[i - 1].value)  # get list of pre-processed prescriptions and a mapping to original string
                common_values = get_common_values(structured, unstructured)  # get common values between structured and unstructured
                if common_values:
                    results_sheet.cell(row_no + 1, i + 1).value = format_list(common_values,
                                                                              structured_true_value_mapping)
                    count_match += 1
                else:
                    results_sheet.cell(row_no + 1, i + 1).value = row[i - 2].value
            else:  # # case for higher priority for unstructured data
                structured, _ = get_values(row[i - 2].value)
                unstructured, unstructured_true_value_mapping = get_values(row[i - 1].value)
                common_values = get_common_values(unstructured, structured)
                if common_values:
                    results_sheet.cell(row_no + 1, i + 1).value = format_list(common_values,
                                                                              unstructured_true_value_mapping)
                    count_match += 1
                else:
                    results_sheet.cell(row_no + 1, i + 1).value = row[i - 1].value
    print("Number of rows with intersecting data between structures and unstructured = ", count_match)
    print("Number of rows with answer retrieved from both the modalities = ", count_populated_cells)
    results_workbook.save(res_file)


# Returns True if only one number each exists in the strings and they are equal. False for all other cases
def get_equality_numeric(str1, str2):
    n_str1 = ""
    i = 0
    while i < len(str1):
        if '0' <= str1[i] <= '9':
            if n_str1 != "":
                return False
            n_str1 += str1[i]
            i += 1
            while i < len(str1) and '0' <= str1[i] <= '9':
                n_str1 += str1[i]
                i += 1
        elif i >= len(str1):
            break
        else:
            i += 1

    n_str2 = ""
    i = 0
    while i < len(str2):
        if '0' <= str2[i] <= '9':
            if n_str2 != "":
                return False
            n_str2 += str2[i]
            i += 1
            while i < len(str2) and '0' <= str2[i] <= '9':
                n_str2 += str2[i]
                i += 1
        elif i >= len(str2):
            break
        else:
            i += 1
    if n_str1 == "" or n_str2 == "":
        return False
    if n_str1 == n_str2:
        return True
    return False


# Function returns strings which satisfy initial substring match and number embedded in string.
# Prioritises representation given in list1
def get_common_values_with_initial_substring_match(list1, list2):
    set1 = set(list1)
    res = list()
    while len(set1) > 0:  # loop checks if shorter sting is contained in the larger string. And if numeric equality exists
        str1 = set1.pop()
        for j in list2:
            if get_equality_numeric(str1, j):
                res.append(str1)
            elif len(str1) >= len(j):
                if j == str1[0:len(j)]:
                    res.append(str1)
            else:
                if str1 == j[0:len(str1)]:
                    res.append(str1)
    return res


def annotate_initial_substring_match(sheet, res_file, column_priority):
    results_workbook = openpyxl.Workbook()
    results_sheet = results_workbook.active
    i = 0
    for row in sheet.iter_rows(max_row=1):
        res_row = list()
        for cell in row:
            res_row.append(cell.value)
            i += 1
        res_row.append("Multimodal_Answer")
        results_sheet.append(res_row)
    count_match = 0
    count_populated_cells = 0
    for row_no, row in enumerate(sheet.iter_rows(min_row=2), start=1):
        for j in range(len(row)):
            results_sheet.cell(row_no + 1, j + 1).value = row[j].value

        if row[i - 2].value is None:
            results_sheet.cell(row_no + 1, i + 1).value = row[i - 1].value
        elif row[i - 1].value is None:
            results_sheet.cell(row_no + 1, i + 1).value = row[i - 2].value
        else:
            count_populated_cells += 1
            if column_priority == 0:
                structured, structured_true_value_mapping = get_values(row[i - 2].value)
                unstructured, _ = get_values(row[i - 1].value)
                common_values = get_common_values_with_initial_substring_match(structured, unstructured)
                if common_values:
                    results_sheet.cell(row_no + 1, i + 1).value = format_list(common_values,
                                                                              structured_true_value_mapping)
                    count_match += 1
                else:
                    results_sheet.cell(row_no + 1, i + 1).value = row[i - 2].value
            else:
                structured, _ = get_values(row[i - 2].value)
                unstructured, unstructured_true_value_mapping = get_values(row[i - 1].value)
                common_values = get_common_values_with_initial_substring_match(unstructured, structured)
                if common_values:
                    results_sheet.cell(row_no + 1, i + 1).value = format_list(common_values,
                                                                              unstructured_true_value_mapping)
                    count_match += 1
                else:
                    results_sheet.cell(row_no + 1, i + 1).value = row[i - 1].value
    print("Number of rows with intersecting data between structures and unstructured = ", count_match)
    print("Number of rows with answer retrieved from both the modalities = ", count_populated_cells)
    results_workbook.save(res_file)


# Function works same as get_common_values_with_initial_substring_match() with added comparison for acronyms
def get_common_values_with_initial_substring_and_acronym_match(list1, list2):
    set1 = set(list1)
    res = list()
    while len(set1) > 0:
        str1 = set1.pop()
        '''if str1 == "TAB" or str1 == "TABLET":
            if "TAB" in list2 or "TABLET" in list2:
                res.append(str1)
            continue
        elif str1 == "CAP" or str1 == "CAPSULE":
            if "CAP" in list2 or "CAPSULE" in list2:
                res.append(str1)
            continue
        elif str1 == "SYR" or str1 == "SYRINGE":
            if "SYR" in list2 or "SYRINGE" in list2:
                res.append(str1)
            continue'''
        if str1 == "IV" or str1 == "INTRAVENOUS":
            if "IV" in list2 or "INTRAVENOUS" in list2:
                res.append(str1)
            continue
        if str1 == "IH" or str1 == "INHALATION":
            if "IH" in list2 or "INHALATION" in list2:
                res.append(str1)
            continue
        for j in list2:
            if get_equality_numeric(str1, j):
                res.append(str1)
            elif len(str1) >= len(j):
                if j == str1[0:len(j)]:
                    res.append(str1)
            else:
                if str1 == j[0:len(str1)]:
                    res.append(str1)
    return res


def annotate_initial_substring_match_with_acronyms(sheet, res_file, column_priority):
    results_workbook = openpyxl.Workbook()
    results_sheet = results_workbook.active
    i = 0
    for row in sheet.iter_rows(max_row=1):
        res_row = list()
        for cell in row:
            res_row.append(cell.value)
            i += 1
        res_row.append("Multimodal_Answer")
        results_sheet.append(res_row)
    count_match = 0
    count_populated_cells = 0
    for row_no, row in enumerate(sheet.iter_rows(min_row=2), start=1):
        for j in range(len(row)):
            results_sheet.cell(row_no + 1, j + 1).value = row[j].value

        if row[i - 2].value is None:
            results_sheet.cell(row_no + 1, i + 1).value = row[i - 1].value
        elif row[i - 1].value is None:
            results_sheet.cell(row_no + 1, i + 1).value = row[i - 2].value
        else:
            count_populated_cells += 1
            if column_priority == 0:
                structured, structured_true_value_mapping = get_values(row[i - 2].value)
                unstructured, _ = get_values(row[i - 1].value)
                common_values = get_common_values_with_initial_substring_and_acronym_match(structured, unstructured)
                if common_values:
                    results_sheet.cell(row_no + 1, i + 1).value = format_list(common_values,
                                                                              structured_true_value_mapping)
                    count_match += 1
                else:
                    results_sheet.cell(row_no + 1, i + 1).value = row[i - 2].value
            else:
                structured, _ = get_values(row[i - 2].value)
                unstructured, unstructured_true_value_mapping = get_values(row[i - 1].value)
                common_values = get_common_values_with_initial_substring_and_acronym_match(unstructured, structured)
                if common_values:
                    results_sheet.cell(row_no + 1, i + 1).value = format_list(common_values,
                                                                              unstructured_true_value_mapping)
                    count_match += 1
                else:
                    results_sheet.cell(row_no + 1, i + 1).value = row[i - 1].value
    print("Number of rows with intersecting data between structures and unstructured = ", count_match)
    print("Number of rows with answer retrieved from both the modalities = ", count_populated_cells)
    results_workbook.save(res_file)


# Function returns prescriptions giving priority to match both prescription and dosage. If none of that case exists,
# returns prescriptions that match disregarding their dosage. Returned dict prioritises dosage from list1.
# Prescription matches follow substring match
def get_common_values_with_prescription_and_dosage(list1, list2):
    dict1 = {}  # stores a dict mapping prescription to dosage in list1
    list1 = list(set(list1))
    for i in list1:
        data = i.split(":")
        dict1.update({data[0]: data[1]})
    dict2 = {}  # stores a dict mapping prescription to dosage in list2
    for i in list2:
        data = i.split(":")
        dict2.update({data[0]: data[1]})
    res_dict1 = {}  # dict that stores matches in both prescription and dosage
    res_dict2 = {}  # dict that stores matches in prescription disregarding dosage
    while len(dict1) > 0:
        row1 = dict1.popitem()
        row2 = dict2.pop(row1[0], "-1")
        if row2 != "-1":
            if get_equality_numeric(row1[1], row2):
                res_dict1.update({row1[0]: row1[1]})
            elif len(row1[1]) >= len(row2):
                if row2 == row1[1][0:len(row2)]:
                    res_dict1.update({row1[0]: row1[1]})
                else:
                    res_dict2.update({row1[0]: row1[1]})
            else:
                if row1[1] == row2[0:len(row1[1])]:
                    res_dict1.update({row1[0]: row1[1]})
                else:
                    res_dict2.update({row1[0]: row1[1]})
    if len(res_dict1) > 0:
        return res_dict1
    return res_dict2


# Function takes a dict of preprocessed strings and dosage, and a dict mapping preprocessed strings to original string.
# Returns string in a format suitable to be written to query file
def format_dict(data_dict, value_mapping):
    res_string = "("
    while len(data_dict) > 0:
        data = data_dict.popitem()
        res_string += value_mapping.pop(data[0]+":"+data[1]) + ","
    res_string = res_string[0: len(res_string) - 1]
    res_string += ')'
    return res_string


def annotate_across_prescription_and_dosage(sheet, res_file, column_priority):
    results_workbook = openpyxl.Workbook()
    results_sheet = results_workbook.active
    i = 0
    for row in sheet.iter_rows(max_row=1):
        res_row = list()
        for cell in row:
            res_row.append(cell.value)
            i += 1
        res_row.append("Multimodal_Answer")
        results_sheet.append(res_row)
    count_match = 0
    count_populated_cells = 0
    for row_no, row in enumerate(sheet.iter_rows(min_row=2), start=1):
        for j in range(len(row)):
            results_sheet.cell(row_no + 1, j + 1).value = row[j].value

        if row[i - 2].value is None:
            results_sheet.cell(row_no + 1, i + 1).value = row[i - 1].value
        elif row[i - 1].value is None:
            results_sheet.cell(row_no + 1, i + 1).value = row[i - 2].value
        else:
            count_populated_cells += 1
            if column_priority == 0:
                structured, structured_true_value_mapping = get_values(row[i - 2].value)
                unstructured, _ = get_values(row[i - 1].value)
                common_values = get_common_values_with_prescription_and_dosage(structured, unstructured)
                if len(common_values) > 0:
                    results_sheet.cell(row_no + 1, i + 1).value = format_dict(common_values,
                                                                              structured_true_value_mapping)
                    count_match += 1
                else:
                    results_sheet.cell(row_no + 1, i + 1).value = row[i - 2].value
            else:
                structured, _ = get_values(row[i - 2].value)
                unstructured, unstructured_true_value_mapping = get_values(row[i - 1].value)
                common_values = get_common_values_with_prescription_and_dosage(unstructured, structured)
                if len(common_values) > 0:
                    results_sheet.cell(row_no + 1, i + 1).value = format_dict(common_values,
                                                                              unstructured_true_value_mapping)
                    count_match += 1
                else:
                    results_sheet.cell(row_no + 1, i + 1).value = row[i - 1].value
    print("Number of rows with intersecting data between structures and unstructured = ", count_match)
    print("Number of rows with answer retrieved from both the modalities = ", count_populated_cells)
    results_workbook.save(res_file)


def main():
    argument_list = sys.argv[1:]
    options = "hi:o:a:p:"
    long_options = ["help", "input_file=", "output_file=", "annotation_flag=", "column_priority="]
    file_path = ""
    output_file_path = ""
    annotation_flag = -1
    column_priority = -1
    try:
        arguments, _ = getopt.getopt(argument_list, options, long_options)
        for currentArgument, currentValue in arguments:
            if currentArgument in ("-h", "--help"):
                print("There are four flags for running this program\n\ninput_file - Path to input xlsx "
                      "file\n\noutput_file - Path to new output xlsx file\n\nannotation_flag - Can take one among "
                      "four values [0/1/2/3]. Pass 0 for exact string match, 1 for initial substring match, "
                      "2 for initial substring match with acronyms, and 3 for matching across prescription and dosage "
                      "(separated by ':')\n\ncolumn_priority - Can take one of two values [0/1]. Pass 0 for a higher "
                      "priority for structured annotation. Pass 1 for a higher priority for unstructured annotation\n")
                return
            elif currentArgument in ("-i", "--input_file"):
                file_path = currentValue
            elif currentArgument in ("-o", "--output_file"):
                output_file_path = currentValue
            elif currentArgument in ("-a", "--annotation_flag"):
                annotation_flag = int(currentValue)
            elif currentArgument in ("-p", "--column_priority"):
                column_priority = int(currentValue)
    except getopt.error as err:
        print(str(err))

    if file_path[-4:len(file_path)].lower() != "xlsx":
        print("input file has to be a xlsx file", file_path, "bla")
        return
    if output_file_path[-4:len(file_path)].lower() != "xlsx":
        print("output file has to be a xlsx file")
        return
    if not(0 <= annotation_flag <= 3):
        print("invalid value for annotation_flag. check help for details")
        return
    if not(0 <= column_priority <= 1):
        print("invalid value for column_priority. check help for details")
        return

    sheet = init(file_path)
    if annotation_flag == 0:
        annotate_exact_match(sheet, output_file_path, column_priority)
    elif annotation_flag == 1:
        annotate_initial_substring_match(sheet, output_file_path, column_priority)
    elif annotation_flag == 2:
        annotate_initial_substring_match_with_acronyms(sheet, output_file_path, column_priority)
    elif annotation_flag == 3:
        annotate_across_prescription_and_dosage(sheet, output_file_path, column_priority)


if __name__ == '__main__':
    main()
